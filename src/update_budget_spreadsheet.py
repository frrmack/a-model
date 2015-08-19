"""
This script is used to download data directly from quickbooks and cache the
results locally.
"""
# NOTE: The quickbooks API is intended for webapps, not for people to download
# their own data. A simple downloading scheme with requests didn't work because
# of some janky ass javascript and iframe bullshit that quickbooks online has.
# Selenium was the best choice.

import urlparse
import time
import os
import datetime
import itertools
import glob

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import gspread
import openpyxl

from datascope import Datascope
import utils

# url for quickbooks homepage
homepage = 'http://qbo.intuit.com'


def get_report_url(params):
    """convenience function for creating report urls"""
    report_url = homepage + '/app/report'
    return report_url + '?' + utils.urlencode(params)

# prepare a bunch of urls for accessing different reports
start_date = datetime.date(2014, 1, 1)
end_date = utils.end_of_last_month()
date_customized_params = (
    ('high_date', utils.qbo_date_str(end_date)),
    ('low_date', utils.qbo_date_str(start_date)),
    ('date_macro', 'custom'),
    ('customized', 'yes'),
)
profit_loss_params = (
    ('rptId', 'reports/ProfitAndLossReport'),
    ('column', 'monthly'),
) + date_customized_params
ar_aging_params = (
    ('rptId', 'AR_AGING'),
)
balance_sheet_params = (
    ('rptId', 'reports/BalanceSheetReport'),
    ('token', 'BAL_SHEET'),
    ('column', 'monthly'),
    ('collapse_subs', 'true'),
) + date_customized_params
unpaid_invoices_params = (
    ('rptId', 'txreports/TxListReport'),
    ('arpaid', '2'),
    ('token', 'INVOICE_LIST'),
) + date_customized_params
profit_loss_url = get_report_url(profit_loss_params)
ar_aging_url = get_report_url(ar_aging_params)
balance_sheet_url = get_report_url(balance_sheet_params)
unpaid_invoices_url = get_report_url(unpaid_invoices_params)
print balance_sheet_url
exit()

# instantiate the datascope object
datascope = Datascope()

# basic filename manipulation
download_dir = datascope.data_root
profit_loss_filename = os.path.join(download_dir, 'profit_loss.xlsx')
ar_aging_filename = os.path.join(download_dir, 'ar_aging.xlsx')
balance_sheet_filename = os.path.join(download_dir, 'balance_sheet.xlsx')
unpaid_invoices_filename = os.path.join(download_dir, 'unpaid_invoices.xlsx')


def open_browser():

    # create a firefox profile to automatically download files (like excel files)
    # without having to approve of the download
    # http://bit.ly/1WeZziv
    profile = webdriver.FirefoxProfile()
    profile.set_preference("browser.download.folderList", 2)
    profile.set_preference("browser.download.manager.showWhenStarting", False)
    profile.set_preference("browser.download.dir", download_dir)
    profile.set_preference(
        "browser.helperApps.neverAsk.saveToDisk",
        ','.join((
            'application/vnd.ms-excel',
            'application/msexcel',
            'application/x-msexcel',
            'application/x-ms-excel',
            'application/x-excel',
            'application/x-dos_ms_excel',
            'application/xls',
            'application/x-xls',
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        ))
    )

    # instantiate a firefox instance and implicitly wait for find_element_* methods
    # for 10 seconds in case content does not immediately appear
    browser = webdriver.Firefox(firefox_profile=profile)
    browser.implicitly_wait(10)
    return browser


# login to quickbooks
def login(browser):
    browser.get(homepage)
    login = browser.find_element_by_name("login")
    login.send_keys(datascope.config.get('quickbooks', 'username'))
    password = browser.find_element_by_name("password")
    password.send_keys(datascope.config.get('quickbooks', 'password'))
    button = browser.find_element_by_id("LoginButton")
    button.click()


def download_report(browser, report_url, xlsx_filename):
    # remove all of the old report*.xlsx crappy filenames that quickbooks
    # creates by default
    report_xlsx_regex = os.path.join(download_dir, 'report*.xlsx')
    for filename in glob.glob(report_xlsx_regex):
        os.remove(filename)

    # go to the P&L page and download the report locally
    browser.get(report_url)
    iframe = browser.find_element_by_tag_name('iframe')
    browser.switch_to_frame(iframe)
    iframe2 = browser.find_element_by_tag_name('iframe')
    browser.switch_to_frame(iframe2)
    xlsx_export = browser.find_element_by_css_selector('option[value=xlsx]')
    xlsx_export.click()
    browser.switch_to_default_content()

    # TODO: find other way of making sure that the file downloaded correctly
    while not glob.glob(report_xlsx_regex):
        time.sleep(1)
    qbo_xlsx_filename = glob.glob(report_xlsx_regex)[0]
    os.rename(qbo_xlsx_filename, xlsx_filename)


def upload_report_to_google(xlsx_filename, gsheet_tab_name):

    # parse the resulting data from xlsx and upload it to a google spreadsheet
    excel_workbook = openpyxl.load_workbook(xlsx_filename)
    excel_worksheet = excel_workbook.active
    pl_dimension = excel_worksheet.calculate_dimension()
    excel_row_list = excel_worksheet.range(pl_dimension)
    excel_cell_list = itertools.chain(*excel_row_list)

    google_workbook = datascope._open_google_workbook()
    google_worksheet = google_workbook.worksheet(gsheet_tab_name)
    google_cell_list = google_worksheet.range(pl_dimension)

    for google_cell, excel_cell in zip(google_cell_list, excel_cell_list):
        if excel_cell.value is None:
            google_cell.value = ''
        else:
            google_cell.value = excel_cell.value
    google_worksheet.update_cells(google_cell_list)


def sync_quickbooks_to_google(browser, report_url, xlsx_filename, tab_name):
    """Convenience method for syncing a report in quickbooks to google spreadsheet
    """
    download_report(browser, report_url, xlsx_filename)
    upload_report_to_google(xlsx_filename, tab_name)


def unpaid_invoices2accounts_receivable(unpaid_invoice_xlsx_filename,
    accounts_receivable_xlsx_filename):

    # instantiate the different workbooks
    invoices_workbook = openpyxl.load_workbook(unpaid_invoice_xlsx_filename)
    invoices_worksheet = invoices_workbook.active
    accounts_receivable_workbook = openpyxl.Workbook()
    accounts_receivable_worksheet = accounts_receivable_workbook.active

    # get the maximum due date of all invoices
    due_date_range = 'G6:G%d' % invoices_worksheet.max_row
    dates = []
    for row in invoices_worksheet.iter_rows(due_date_range):
        for cell in row:
            if cell.value is not None:
                date = utils.qbo_date(cell.value)
                dates.append(utils.end_of_month(date))
    max_date = max(dates)

    # add a row of all months during which we expect to get paid
    month2column = {}
    row_index = 1
    for i, month in enumerate(utils.iter_end_of_months(start_date, max_date)):
        column = i + 2
        month2column[month] = column = i+2
        accounts_receivable_worksheet.cell(
            column=column, row=row_index, value=month,
        )

    # iterate over all of the invoices and add them to the
    data_range = 'B6:I%d' % invoices_worksheet.max_row
    row_index += 2
    client_name2row_index = {}
    for row in invoices_worksheet.iter_rows(data_range):
        client_name = row[3].value
        if client_name is None:
            break
        due_date = utils.qbo_date(row[5].value)
        month_due = utils.end_of_month(due_date)
        unpaid_balance = row[7].value
        if client_name not in client_name2row_index:
            client_name2row_index[client_name] = row_index
            accounts_receivable_worksheet.cell(
                column=1, row=row_index, value=client_name,
            )
            row_index += 1
        accounts_receivable_worksheet.cell(
            column=month2column[month_due],
            row=client_name2row_index[client_name],
            value=unpaid_balance
        )

    # create a total row
    accounts_receivable_worksheet.cell(column=1, row=2, value="TOTAL")
    for column in month2column.itervalues():
        col = openpyxl.utils.get_column_letter(column)
        accounts_receivable_worksheet.cell(
            column=column,
            row=2,
            value="=SUM(%s3:%s%d)" % (col, col, row_index-1),
        )

    # save the accounts receivable report
    accounts_receivable_workbook.save(accounts_receivable_xlsx_filename)



if __name__=='__main__':
    browser = open_browser()
    login(browser)

    # sync quickbooks report to google spreadsheet
    sync_quickbooks_to_google(
        browser, profit_loss_url, profit_loss_filename, "P&L",
    )
    sync_quickbooks_to_google(
        browser, ar_aging_url, ar_aging_filename, "A/R Aging",
    )
    sync_quickbooks_to_google(
        browser, balance_sheet_url, balance_sheet_filename, "Balance Sheet",
    )

    # download the unpaid invoices report from quickbooks, which is used in
    # revenue projection simulations later
    download_report(browser, unpaid_invoice_url, unpaid_invoice_xlsx_filename)
    # unpaid_invoices2accounts_receivable(
    #     unpaid_invoice_xlsx_filename,
    #     accounts_receivable_xlsx_filename,
    # )

    # close the browser
    browser.close()
